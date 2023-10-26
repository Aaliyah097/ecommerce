import sys
import os
import django
import openpyxl
import requests


sys.path.append(os.path.join(os.path.dirname(__file__), 'ecommerce'))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "ecommerce.settings")
django.setup()

from django.db import transaction
from catalog.models import *
from bs4 import BeautifulSoup

from catalog import utils


def import_routers_switch(page: str):
    wb = openpyxl.load_workbook(r"C:\Users\Aaliyah\dev\gamma\routers-switch.com-cards.xlsx")
    sheet = wb[page]
    source = 'www.router-switch.com'
    currency = Currencies.objects.get(name='USD')
    for row_number in range(2, sheet.max_row + 1):
        product = {
            'source': source,
            'part_number': sheet.cell(row_number, 1).value,
            'category': sheet.cell(row_number, 2).value,
            'brand': sheet.cell(row_number, 3).value,
            'sub_cat1': sheet.cell(row_number, 4).value or None,
            'sub_cat2': sheet.cell(row_number, 5).value,
            'sub_cat3': sheet.cell(row_number, 6).value,
            'series': sheet.cell(row_number, 7).value,
            'link': sheet.cell(row_number, 8).value,
            'price': sheet.cell(row_number, 9).value,
            'name': sheet.cell(row_number, 10).value,
            'picture_link': sheet.cell(row_number, 11).value,
            'specs': sheet.cell(row_number, 12).value or None
        }

        brand = Brands.objects.get(name=product['brand'])
        try:
            category = Categories.objects.get(name=product['sub_cat1'])
        except Categories.DoesNotExist:
            category = Categories.objects.get(slug=product['category'].lower())

        new_product = Products(
            name=product['name'],
            category=category,
            part_number=product['part_number'],
            brand=brand,
            price=product['price'],
            currency=currency,
            description=product['specs'],
            source=source,
            source_link=product['link'],
            series=product['series'],
            image_link=product['picture_link']
        )
        try:
            new_product = Products.objects.get(part_number=product['part_number'], brand=brand)
        except Products.DoesNotExist:
            new_product.save()


def iter_products(sheet) -> dict:
    for row_number in range(2, sheet.max_row + 1):
        yield {
            'part_number': sheet.cell(row_number, 1).value,
            'category': sheet.cell(row_number, 2).value,
            'brand': sheet.cell(row_number, 3).value,
            'sub_cat1': sheet.cell(row_number, 4).value,
            'sub_cat2': sheet.cell(row_number, 5).value,
            'sub_cat3': sheet.cell(row_number, 6).value,
            'series': sheet.cell(row_number, 7).value,
            'link': sheet.cell(row_number, 8).value,
            'price': sheet.cell(row_number, 9).value,
            'name': sheet.cell(row_number, 10).value,
            'picture_link': sheet.cell(row_number, 11).value,
            'specs': sheet.cell(row_number, 12).value or None
        }


def update_products(page: str, source: str):
    wb = openpyxl.load_workbook(r"C:\Users\Aaliyah\dev\gamma\routers-switch.com-cards.xlsx")
    sheet = wb[page]
    currency = Currencies.objects.get(name='USD')
    for product in iter_products(sheet):
        product['source'] = source

        brand = Brands.objects.get(name=product['brand'])
        try:
            category = Categories.objects.get(name=product['sub_cat1'])
        except Categories.DoesNotExist:
            category = Categories.objects.get(name=product['category'])

        try:
            ex_product = Products.objects.get(part_number=product['part_number'], brand=brand)
            ex_product.category = category
            ex_product.save()
        except Products.DoesNotExist:
            continue


if __name__ == '__main__':
    # update_products('routers', 'www.router-switch.com')
    import_routers_switch('routers')
